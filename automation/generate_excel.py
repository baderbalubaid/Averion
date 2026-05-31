import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
import os
import json
import subprocess
from datetime import datetime, timedelta
from dotenv import load_dotenv

load_dotenv()

DB_CONFIG = {
   'host': os.getenv('DB_HOST', 'localhost'),
   'port': os.getenv('DB_PORT', 5432),
   'dbname': os.getenv('DB_NAME', 'averion'),
   'user': os.getenv('DB_USER', 'averion'),
   'password': os.getenv('DB_PASSWORD')
}

# ═══════════════════════════════
# COLOR STANDARDS
# ═══════════════════════════════
C_HEADER_BG   = 'FF0E0E1C'  # Dark header
C_HEADER_FG   = 'FFFFFFFF'  # White text
C_PROFIT      = 'FF10D98A'  # Green
C_LOSS        = 'FFF4645F'  # Red
C_WARNING     = 'FFFFA500'  # Amber
C_INFO        = 'FF38BDF8'  # Blue
C_BENCHMARK   = 'FF888888'  # Gray
C_ELIMINATED  = 'FF444444'  # Dark gray
C_ALT_ROW     = 'FF1A1A2E'  # Alternate row

def style_header(ws, row, columns):
   for col, title in enumerate(columns, 1):
       cell = ws.cell(row=row, column=col, value=title)
       cell.font = Font(bold=True, color=C_HEADER_FG, name='Calibri', size=10)
       cell.fill = PatternFill(fill_type='solid', fgColor=C_HEADER_BG)
       cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
       ws.column_dimensions[get_column_letter(col)].width = max(len(title) + 4, 12)

def freeze_and_filter(ws, freeze_cell, filter_range):
   ws.freeze_panes = freeze_cell
   ws.auto_filter.ref = filter_range

def generate_report():
   now = datetime.utcnow()
   today = now.date()
   yesterday = today - timedelta(days=1)

   try:
       conn = psycopg2.connect(**DB_CONFIG)
       cur = conn.cursor()
       db_connected = True
   except:
       db_connected = False
       conn = None
       cur = None

   wb = openpyxl.Workbook()
   wb.remove(wb.active)

   # ═══════════════════════════════
   # SHEET 1 — Daily Summary
   # ═══════════════════════════════
   ws1 = wb.create_sheet('01_Daily_Summary')
   ws1.column_dimensions['A'].width = 35
   ws1.column_dimensions['B'].width = 25

   title = ws1.cell(row=1, column=1, value=f'Averion Daily Report — {today}')
   title.font = Font(bold=True, size=16, color=C_HEADER_FG, name='Calibri')
   title.fill = PatternFill(fill_type='solid', fgColor=C_HEADER_BG)
   ws1.merge_cells('A1:B1')

   metrics = [
       ('Report Date', str(today)),
       ('Generated At', now.strftime('%Y-%m-%d %H:%M UTC')),
       ('', ''),
   ]

   if db_connected:
       cur.execute("SELECT COUNT(*) FROM positions WHERE status='open' AND is_paper=FALSE")
       live_open = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM positions WHERE status='open' AND is_paper=TRUE")
       paper_open = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM positions WHERE status='closed' AND DATE(closed_at)=%s", (yesterday,))
       closed_yest = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM bots WHERE status='active'")
       active_bots = cur.fetchone()[0]
       cur.execute("SELECT COALESCE(SUM(balance_usdt),0) FROM reserve_wallets")
       reserve = float(cur.fetchone()[0])
       cur.execute("SELECT accumulated_fees_usdt FROM owner_balance LIMIT 1")
       row = cur.fetchone()
       owner_bal = float(row[0]) if row else 0

       metrics += [
           ('Open Positions Live', live_open),
           ('Open Positions Paper', paper_open),
           ('Total Open', live_open + paper_open),
           ('Closed Yesterday', closed_yest),
           ('', ''),
           ('Active Bots', active_bots),
           ('Reserve Wallet Balance', f'${reserve:.2f}'),
           ('Owner Wallet Balance', f'${owner_bal:.2f}'),
       ]
   else:
       metrics.append(('Database', 'Not connected'))

   for i, (label, value) in enumerate(metrics, 2):
       ws1.cell(row=i, column=1, value=label).font = Font(bold=True, name='Calibri')
       ws1.cell(row=i, column=2, value=value)

   # ═══════════════════════════════
   # SHEET 2 — Open Positions
   # ═══════════════════════════════
   ws2 = wb.create_sheet('02_Open_Positions')
   cols2 = [
       'Position ID', 'Coin', 'Pair', 'Exchange', 'Bot ID',
       'Method', 'Direction', 'Entry Date', 'Days Open',
       'Avg Cost', 'DCA Level', 'Total Invested',
       'Next DCA Trigger', 'TP Price', 'Category',
       'Status', 'Paper/Live', 'Last Update'
   ]
   style_header(ws2, 1, cols2)
   freeze_and_filter(ws2, 'A2', f'A1:{get_column_letter(len(cols2))}1')

   if db_connected:
       cur.execute("""
           SELECT p.id, p.coin, p.coin, e.exchange,
                  b.id, b.method, p.direction,
                  p.opened_at,
                  EXTRACT(DAY FROM NOW()-p.opened_at),
                  p.avg_cost, p.dca_count, p.total_invested,
                  p.last_buy_price, p.avg_cost,
                  p.category, p.status,
                  CASE WHEN p.is_paper THEN 'Paper' ELSE 'Live' END,
                  NOW()
           FROM positions p
           JOIN exchanges e ON e.id = p.exchange_id
           JOIN bots b ON b.id = p.bot_id
           WHERE p.status = 'open'
           ORDER BY p.total_invested DESC
       """)
       for row_idx, row in enumerate(cur.fetchall(), 2):
           for col_idx, val in enumerate(row, 1):
               cell = ws2.cell(row=row_idx, column=col_idx,
                              value=str(val) if isinstance(val, datetime) else val)
               if row_idx % 2 == 0:
                   cell.fill = PatternFill(fill_type='solid', fgColor=C_ALT_ROW)

   # ═══════════════════════════════
   # SHEET 3 — Closed Yesterday
   # ═══════════════════════════════
   ws3 = wb.create_sheet('03_Closed_Yesterday')
   cols3 = [
       'Position ID', 'Coin', 'Exchange', 'Bot ID', 'Method',
       'Entry Method', 'Direction', 'Entry Date', 'Exit Date',
       'Hold Hours', 'Avg Cost', 'Exit Price', 'DCA Levels Used',
       'Total Invested', 'Gross Profit', 'Exchange Fees',
       'Net Profit', 'Performance Fee', 'Final Net Profit',
       'ROI %', 'Exit Reason', 'Paper/Live'
   ]
   style_header(ws3, 1, cols3)
   freeze_and_filter(ws3, 'A2', f'A1:{get_column_letter(len(cols3))}1')

   if db_connected:
       cur.execute("""
           SELECT p.id, p.coin, e.exchange, b.id, b.method,
                  p.entry_method, p.direction,
                  p.opened_at, p.closed_at,
                  ROUND(EXTRACT(EPOCH FROM (p.closed_at-p.opened_at))/3600, 1),
                  p.avg_cost, p.avg_cost,
                  p.dca_count, p.total_invested,
                  0, 0, 0, 0, 0, 0,
                  p.close_reason,
                  CASE WHEN p.is_paper THEN 'Paper' ELSE 'Live' END
           FROM positions p
           JOIN exchanges e ON e.id = p.exchange_id
           JOIN bots b ON b.id = p.bot_id
           WHERE p.status = 'closed'
           AND DATE(p.closed_at) = %s
           ORDER BY p.closed_at DESC
       """, (yesterday,))
       for row_idx, row in enumerate(cur.fetchall(), 2):
           for col_idx, val in enumerate(row, 1):
               cell = ws3.cell(row=row_idx, column=col_idx,
                              value=str(val) if isinstance(val, datetime) else val)

   # ═══════════════════════════════
   # SHEET 4 — Research Rankings
   # ═══════════════════════════════
   ws4 = wb.create_sheet('04_Research_Rankings')
   cols4 = [
       'Rank', 'Bot ID', 'Method', 'Status',
       'Total Trades', 'Wins', 'Losses', 'Win Rate %',
       'Total Profit', 'Avg Win', 'Avg Loss',
       'Max Drawdown', 'Avg Hold Hours', 'Recovery Speed',
       'vs BTC Hold', 'vs Simple DCA', 'vs Random Entry',
       'Promotion Score', 'Sample Size Rating', 'Confidence Level'
   ]
   style_header(ws4, 1, cols4)
   freeze_and_filter(ws4, 'A2', f'A1:{get_column_letter(len(cols4))}1')

   # ═══════════════════════════════
   # SHEET 5 — Bot Parameters
   # ═══════════════════════════════
   ws5 = wb.create_sheet('05_Bot_Parameters')
   cols5 = ['Bot ID', 'Method', 'Parameter Name', 'Parameter Value']
   style_header(ws5, 1, cols5)
   freeze_and_filter(ws5, 'A2', f'A1:{get_column_letter(len(cols5))}1')

   try:
       with open('setup/research_bots.json') as f:
           config = json.load(f)
       row_idx = 2
       for method_id, method_data in config['methods'].items():
           for bot in method_data['bots']:
               for param, value in bot.items():
                   if param != 'id':
                       ws5.cell(row=row_idx, column=1, value=bot['id'])
                       ws5.cell(row=row_idx, column=2, value=method_data['name'])
                       ws5.cell(row=row_idx, column=3, value=param)
                       ws5.cell(row=row_idx, column=4, value=value)
                       row_idx += 1
   except:
       ws5.cell(row=2, column=1, value='research_bots.json not found')

   # ═══════════════════════════════
   # SHEET 6 — Trades Ledger
   # ═══════════════════════════════
   ws6 = wb.create_sheet('06_Trades_Ledger')
   cols6 = [
       'Order ID', 'Position ID', 'Timestamp', 'Exchange',
       'Coin', 'Pair', 'Side', 'Order Type',
       'Price', 'Quantity', 'Order Value',
       'Fee', 'Fee Currency', 'Bot ID', 'Paper/Live'
   ]
   style_header(ws6, 1, cols6)
   freeze_and_filter(ws6, 'A2', f'A1:{get_column_letter(len(cols6))}1')

   if db_connected:
       cur.execute("""
           SELECT t.id, t.position_id, t.timestamp,
                  e.exchange, t.coin, t.coin,
                  t.side, t.order_type,
                  t.price, t.quantity, t.usdt_amount,
                  t.exchange_fee, t.fee_currency,
                  t.bot_id,
                  CASE WHEN t.is_paper THEN 'Paper' ELSE 'Live' END
           FROM trades t
           JOIN exchanges e ON e.id = t.exchange_id
           WHERE DATE(t.timestamp) = %s
           ORDER BY t.timestamp DESC
       """, (yesterday,))
       for row_idx, row in enumerate(cur.fetchall(), 2):
           for col_idx, val in enumerate(row, 1):
               cell = ws6.cell(row=row_idx, column=col_idx,
                              value=str(val) if isinstance(val, datetime) else val)
               if row_idx % 2 == 0:
                   cell.fill = PatternFill(fill_type='solid', fgColor=C_ALT_ROW)

   # ═══════════════════════════════
   # SHEET 7 — Platform Stats
   # ═══════════════════════════════
   ws7 = wb.create_sheet('07_Platform_Stats')
   ws7.column_dimensions['A'].width = 35
   ws7.column_dimensions['B'].width = 20

   style_header(ws7, 1, ['Metric', 'Value'])

   platform_metrics = [
       ('Report Date', str(today)),
       ('Active Bots', 0),
       ('Live Bots', 0),
       ('Paper Bots', 0),
       ('Open Positions', 0),
       ('Reserve Wallet Total', '$0.00'),
       ('Owner Wallet Balance', '$0.00'),
       ('Outstanding Fee Debt', '$0.00'),
       ('Lifetime Trades', 0),
       ('Lifetime Positions Closed', 0),
   ]

   if db_connected:
       cur.execute("SELECT COUNT(*) FROM bots WHERE status='active' AND is_paper=FALSE")
       live_bots = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM bots WHERE status='active' AND is_paper=TRUE")
       paper_bots = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM positions WHERE status='open'")
       open_pos = cur.fetchone()[0]
       cur.execute("SELECT COALESCE(SUM(balance_usdt),0) FROM reserve_wallets")
       reserve_total = float(cur.fetchone()[0])
       cur.execute("SELECT accumulated_fees_usdt FROM owner_balance LIMIT 1")
       row = cur.fetchone()
       owner = float(row[0]) if row else 0
       cur.execute("SELECT COALESCE(SUM(amount_usdt),0) FROM fee_debt WHERE paid_at IS NULL")
       debt = float(cur.fetchone()[0])
       cur.execute("SELECT COUNT(*) FROM trades")
       total_trades = cur.fetchone()[0]
       cur.execute("SELECT COUNT(*) FROM positions WHERE status='closed'")
       total_closed = cur.fetchone()[0]

       platform_metrics = [
           ('Report Date', str(today)),
           ('Active Bots', live_bots + paper_bots),
           ('Live Bots', live_bots),
           ('Paper Bots', paper_bots),
           ('Open Positions', open_pos),
           ('Reserve Wallet Total', f'${reserve_total:.2f}'),
           ('Owner Wallet Balance', f'${owner:.2f}'),
           ('Outstanding Fee Debt', f'${debt:.2f}'),
           ('Lifetime Trades', total_trades),
           ('Lifetime Positions Closed', total_closed),
       ]

   for i, (label, value) in enumerate(platform_metrics, 2):
       ws7.cell(row=i, column=1, value=label).font = Font(bold=True)
       ws7.cell(row=i, column=2, value=value)

   # ═══════════════════════════════
   # SHEET 8 — Data Dictionary
   # ═══════════════════════════════
   ws8 = wb.create_sheet('08_Data_Dictionary')
   style_header(ws8, 1, ['Field Name', 'Sheet', 'Meaning', 'Example'])
   ws8.column_dimensions['A'].width = 25
   ws8.column_dimensions['B'].width = 20
   ws8.column_dimensions['C'].width = 45
   ws8.column_dimensions['D'].width = 20

   dictionary = [
       ('Position ID', 'All', 'Unique position identifier in database', '1042'),
       ('Coin', 'All', 'Trading coin symbol', 'RVN'),
       ('Pair', 'All', 'Full trading pair', 'RVN/USDT'),
       ('Direction', 'All', 'Long (buy dips) or Short (sell rises)', 'long'),
       ('Avg Cost', 'Open/Closed', 'Weighted average buy price', '0.0934'),
       ('DCA Level', 'Open', 'Number of DCA buys executed so far', '3'),
       ('Total Invested', 'All', 'Total USDT spent on this position', '47.50'),
       ('Next DCA Trigger', 'Open', 'Price at which next DCA will fire', '0.0867'),
       ('TP Price', 'Open', 'Take profit target price', '0.0982'),
       ('Category', 'All', 'Market cap category (Mega/Large/Mid/Small/Micro)', 'Small Cap'),
       ('ST Flag', 'All', 'Exchange suspended trading flag', 'FALSE'),
       ('Promotion Score', 'Rankings', 'Formula: WR^0.30 x AP^0.20 x RS^0.15 x DD^0.35', '0.742'),
       ('Win Rate %', 'Rankings', 'Percentage of positions closed profitably', '68.5'),
       ('Recovery Speed', 'Rankings', 'Average hours from entry to TP exit', '42.3'),
       ('Max Drawdown', 'Rankings', 'Largest unrealized loss during position', '-$12.40'),
       ('Confidence Level', 'Rankings', 'Low < 30 trades · Medium 30-100 · High > 100', 'Medium'),
       ('ROI %', 'Closed', 'Final net profit divided by total invested', '4.2%'),
       ('Exchange Fee', 'Ledger', 'Actual fee charged by exchange per order', '0.0023'),
       ('Performance Fee', 'Closed', '20% of net profit after exchange fees', '0.89'),
       ('Exit Reason', 'Closed', 'TP = take profit · ST = suspended · Manual', 'TP'),
       ('Paper/Live', 'All', 'Paper = simulated · Live = real money', 'Paper'),
       ('Hold Hours', 'Closed', 'Total hours position was open', '38.5'),
       ('Sample Size Rating', 'Rankings', 'Statistical reliability of score', 'Low'),
       ('Bot ID', 'All', 'Unique bot identifier', 'E1-3'),
   ]

   for i, row in enumerate(dictionary, 2):
       for j, val in enumerate(row, 1):
           ws8.cell(row=i, column=j, value=val)

   # ═══════════════════════════════
   # SHEET 99 — Metadata
   # ═══════════════════════════════
   ws99 = wb.create_sheet('99_Metadata')
   style_header(ws99, 1, ['Key', 'Value'])
   ws99.column_dimensions['A'].width = 30
   ws99.column_dimensions['B'].width = 50

   try:
       git_hash = subprocess.check_output(
           ['git', 'rev-parse', '--short', 'HEAD'],
           stderr=subprocess.DEVNULL
       ).decode().strip()
   except:
       git_hash = 'unknown'

   metadata = [
       ('Report Version', '1.0'),
       ('Schema Version', '1.0'),
       ('Generation Timestamp', now.strftime('%Y-%m-%d %H:%M:%S UTC')),
       ('Git Commit Hash', git_hash),
       ('Report Date', str(today)),
       ('Database Connected', str(db_connected)),
       ('Python Script', 'automation/generate_excel.py'),
       ('Averion Version', 'Phase 4'),
   ]

   for i, (key, value) in enumerate(metadata, 2):
       ws99.cell(row=i, column=1, value=key).font = Font(bold=True)
       ws99.cell(row=i, column=2, value=value)

   if conn:
       conn.close()

   # Save file
   os.makedirs('reports', exist_ok=True)
   filename = f'reports/averion_report_{today}.xlsx'
   wb.save(filename)
   print(f'✅ Excel report saved: {filename}')
   print(f'✅ Sheets: 01_Daily_Summary · 02_Open_Positions · 03_Closed_Yesterday')
   print(f'✅ Sheets: 04_Research_Rankings · 05_Bot_Parameters · 06_Trades_Ledger')
   print(f'✅ Sheets: 07_Platform_Stats · 08_Data_Dictionary · 99_Metadata')
   return filename

if __name__ == '__main__':
   generate_report()
